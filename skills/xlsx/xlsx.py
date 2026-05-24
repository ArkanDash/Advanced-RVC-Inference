#!/usr/bin/env python3
"""
xlsx.py — Unified Excel Quality Assurance & Manipulation Tool

Commands:
  recalc       <xlsx> [timeout]       Recalculate formulas via LibreOffice + error scan
  audit      <xlsx>                 Formula error + zero-value + implicit array detection
  scan     <xlsx>                 Reference anomaly detection
  inspect      <xlsx> [--pretty]      Structure analysis → JSON
  pivot        <in> <out> [options]   PivotTable with optional chart
  chart-verify <xlsx>                 Verify chart data content
  validate     <xlsx>                 Structural validation (forbidden funcs, schema)

Usage:
  python3 xlsx.py <command> [args...]
  python3 xlsx.py --help
"""

from __future__ import annotations

import argparse
import json
import os
import platform
import re
import shutil
import subprocess
import sys
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string
except ImportError:
    print("Error: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════
#  Section 0: Command registry + shared constants & helpers
# ═══════════════════════════════════════════════════════════════

_COMMANDS: Dict[str, Callable] = {}


def cmd(name: str):
    """Decorator that registers a function as a CLI command."""
    def decorator(fn: Callable) -> Callable:
        _COMMANDS[name] = fn
        return fn
    return decorator


# --------------- constants ---------------

EXCEL_ERRORS = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}

# Implicit array formula patterns that work in LibreOffice but fail in MS Excel
IMPLICIT_ARRAY_PATTERNS = [
    (re.compile(r'MATCH\s*\(\s*TRUE\s*\(\s*\)', re.IGNORECASE),
     "MATCH(TRUE(), ...) requires CSE in Excel. Use SUMPRODUCT or helper column."),
    (re.compile(r'MATCH\s*\(\s*TRUE\s*,', re.IGNORECASE),
     "MATCH(TRUE, ...) with comparison range requires CSE. Use SUMPRODUCT."),
    (re.compile(r'MATCH\s*\([^,]+[<>=!]+[^,]+,', re.IGNORECASE),
     "MATCH with inline comparison requires CSE. Use SUMPRODUCT or helper column."),
]

FORBIDDEN_FUNCTIONS = {
    "FILTER", "UNIQUE", "SORT", "SORTBY", "XLOOKUP", "XMATCH",
    "SEQUENCE", "LET", "LAMBDA", "RANDARRAY",
    "ARRAYFORMULA", "QUERY", "IMPORTRANGE",
}

# [Fix ①] Pattern to detect valid formula content (function calls or cell references)
# A real formula must contain at least one of:
#   - Function call: ALPHA_CHARS( e.g. SUM(, IF(, VLOOKUP(
#   - Cell reference: $?[A-Z]{1,3}$?\d+  e.g. A1, $B$5, $A$1:$A$10
_VALID_FORMULA_PATTERN = re.compile(
    r'[A-Z]{2,}\s*\('           # function call (2+ uppercase letters followed by parenthesis)
    r'|'
    r'\$?[A-Z]{1,3}\$?\d+'      # cell reference like A1, $B$5
    ,
    re.IGNORECASE,
)

# [Fix ②] Pattern to detect external file references in formulas
# Matches [filename.xlsx]SheetName! or [filename.xls]SheetName! etc.
_EXT_FILE_REF_PATTERN = re.compile(r"\[([^\]]+\.(xlsx?|xlsm|xlsb|csv))\]", re.IGNORECASE)


# --------------- helpers ---------------

def cell_ref(sheet_name: str, cell) -> str:
    return f"{sheet_name}!{cell.coordinate}"


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def parse_range(range_str: str) -> Tuple[Optional[str], int, int, int, int]:
    """Parse 'Sheet!A1:F100' into (sheet, min_col, min_row, max_col, max_row)."""
    if "!" in range_str:
        sheet, rng = range_str.rsplit("!", 1)
        sheet = sheet.strip("'\"")
    else:
        sheet = None
        rng = range_str
    parts = rng.split(":")
    if len(parts) == 2:
        c1, r1 = coordinate_from_string(parts[0])
        c2, r2 = coordinate_from_string(parts[1])
        return sheet, column_index_from_string(c1), r1, column_index_from_string(c2), r2
    else:
        c1, r1 = coordinate_from_string(parts[0])
        return sheet, column_index_from_string(c1), r1, column_index_from_string(c1), r1


# ═══════════════════════════════════════════════════════════════
#  Section 1: recalc — LibreOffice recalculation + error scan
# ═══════════════════════════════════════════════════════════════

def _find_soffice() -> Optional[str]:
    """Locate soffice binary across macOS / Linux / Windows.

    Search order:
      1. PATH (shutil.which)
      2. Platform-specific well-known locations
    Returns the absolute path, or None if not found.
    """
    # 1. Check PATH first (works on all platforms if user configured it)
    found = shutil.which("soffice")
    if found:
        return found

    # 2. Platform-specific search
    system = platform.system()

    if system == "Darwin":
        candidates = [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            os.path.expanduser("~/Applications/LibreOffice.app/Contents/MacOS/soffice"),
        ]
    elif system == "Linux":
        candidates = [
            "/usr/bin/soffice",
            "/usr/local/bin/soffice",
            "/usr/lib/libreoffice/program/soffice",
            "/opt/libreoffice/program/soffice",
            "/snap/bin/libreoffice.soffice",        # Snap package
            "/var/lib/flatpak/exports/bin/org.libreoffice.LibreOffice",  # Flatpak
        ]
    elif system == "Windows":
        candidates = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
    else:
        candidates = []

    for path in candidates:
        if os.path.isfile(path):
            return path

    return None


def _setup_libreoffice_macro() -> bool:
    """Setup LibreOffice macro for recalculation if not already configured."""
    if platform.system() == "Darwin":
        macro_dir = os.path.expanduser(
            "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
        )
    else:
        macro_dir = os.path.expanduser(
            "~/.config/libreoffice/4/user/basic/Standard"
        )

    macro_file = os.path.join(macro_dir, "Module1.xba")

    if os.path.exists(macro_file):
        with open(macro_file, "r") as f:
            if "RecalculateAndSave" in f.read():
                return True

    if not os.path.exists(macro_dir):
        soffice_bin = _find_soffice()
        if soffice_bin:
            subprocess.run(
                [soffice_bin, "--headless", "--terminate_after_init"],
                capture_output=True, timeout=10,
            )
        os.makedirs(macro_dir, exist_ok=True)

    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''

    try:
        with open(macro_file, "w") as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def _libreoffice_recalc(filename: str, timeout: int = 30) -> Dict[str, Any]:
    """
    Recalculate formulas in an Excel file via LibreOffice,
    then scan ALL cells for errors.

    Returns a dict with status, error counts, and locations.
    """
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not _setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    # Locate soffice binary (cross-platform)
    soffice_bin = _find_soffice()
    if not soffice_bin:
        return {"error": "LibreOffice not found. Install it and ensure soffice is in PATH."}

    lo_cmd: List[str] = [
        soffice_bin, "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave"
        "?language=Basic&location=application",
        abs_path,
    ]

    # Wrap with a timeout binary when available
    if platform.system() != "Windows":
        timeout_bin: Optional[str] = None
        if platform.system() == "Linux":
            timeout_bin = "timeout"
        elif platform.system() == "Darwin":
            try:
                subprocess.run(
                    ["gtimeout", "--version"],
                    capture_output=True, timeout=1, check=False,
                )
                timeout_bin = "gtimeout"
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass
        if timeout_bin:
            lo_cmd = [timeout_bin, str(timeout)] + lo_cmd

    result = subprocess.run(lo_cmd, capture_output=True, text=True)

    if result.returncode != 0 and result.returncode != 124:
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        else:
            return {"error": error_msg}

    # Scan recalculated file for Excel errors
    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors_list = [
            "#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"
        ]
        error_details: Dict[str, List[str]] = {err: [] for err in excel_errors_list}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors_list:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break
        wb.close()

        out: Dict[str, Any] = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                out["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        # Count formulas for context
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb_formulas.close()

        out["total_formulas"] = formula_count
        return out

    except Exception as e:
        return {"error": str(e)}


@cmd("recalc")
def cmd_recalc(argv: Sequence[str]) -> int:
    """Recalculate formulas via LibreOffice and report errors."""
    parser = argparse.ArgumentParser(prog="xlsx.py recalc",
                                     description="Recalculate Excel formulas via LibreOffice")
    parser.add_argument("file", help="Excel file path")
    parser.add_argument("timeout", nargs="?", type=int, default=30,
                        help="Timeout in seconds (default: 30)")
    args = parser.parse_args(argv)

    result = _libreoffice_recalc(args.file, args.timeout)
    print(json.dumps(result, indent=2, ensure_ascii=False))
    if "error" in result:
        return 1
    return 0 if result.get("total_errors", 0) == 0 else 1


# ═══════════════════════════════════════════════════════════════
#  Section 2: audit — Formula error + zero-value + implicit array
# ═══════════════════════════════════════════════════════════════

def _run_libreoffice_recalc_best_effort(filepath: str) -> None:
    """Attempt LibreOffice recalc (best-effort, swallow errors)."""
    try:
        _libreoffice_recalc(filepath, timeout=30)
    except Exception:
        pass


@cmd("audit")
def cmd_audit(argv: Sequence[str]) -> int:
    """Detect formula errors, zero-value formulas, and implicit array formulas."""
    parser = argparse.ArgumentParser(prog="xlsx.py audit",
                                     description="Formula error scan + zero-value + implicit array detection")
    parser.add_argument("file", help="Excel file path")
    args = parser.parse_args(argv)

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        return 1

    # Best-effort LibreOffice recalculation first
    _run_libreoffice_recalc_best_effort(str(path))

    wb_data = load_workbook(str(path), data_only=True)
    wb_form = load_workbook(str(path), data_only=False)

    errors: List[Dict[str, Any]] = []
    zero_values: List[Dict[str, str]] = []
    implicit_arrays: List[Dict[str, str]] = []
    total_formulas = 0

    for sname in wb_form.sheetnames:
        ws_d = wb_data[sname]
        ws_f = wb_form[sname]
        for row_d, row_f in zip(ws_d.iter_rows(), ws_f.iter_rows()):
            for cd, cf in zip(row_d, row_f):
                fval = cf.value
                if not is_formula(fval):
                    continue
                total_formulas += 1
                fstr = str(fval)

                # Check for formula errors in calculated value
                dval = cd.value
                if dval is not None and isinstance(dval, str):
                    for err in EXCEL_ERRORS:
                        if err in dval:
                            errors.append({
                                "cell": cell_ref(sname, cd),
                                "error": err,
                                "formula": fstr[:120],
                            })
                            break

                # Check for zero values (potential reference errors)
                if isinstance(dval, (int, float)) and dval == 0:
                    zero_values.append({
                        "cell": cell_ref(sname, cd),
                        "formula": fstr[:120],
                    })

                # Check for implicit array formula patterns
                for pattern, msg in IMPLICIT_ARRAY_PATTERNS:
                    if pattern.search(fstr):
                        implicit_arrays.append({
                            "cell": cell_ref(sname, cf),
                            "formula": fstr[:120],
                            "issue": msg,
                        })
                        break

    wb_data.close()
    wb_form.close()

    result: Dict[str, Any] = {
        "total_formulas": total_formulas,
        "error_count": len(errors),
        "zero_value_count": len(zero_values),
        "implicit_array_count": len(implicit_arrays),
    }
    if errors:
        result["errors"] = errors[:50]
    if zero_values:
        result["zero_values"] = zero_values[:30]
    if implicit_arrays:
        result["implicit_arrays"] = implicit_arrays[:20]

    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 1 if errors else 0


# ═══════════════════════════════════════════════════════════════
#  Section 3: scan — Reference anomaly detection
# ═══════════════════════════════════════════════════════════════

@cmd("scan")
def cmd_scan(argv: Sequence[str]) -> int:
    """Detect reference anomalies in formulas."""
    parser = argparse.ArgumentParser(prog="xlsx.py scan",
                                     description="Reference anomaly detection")
    parser.add_argument("file", help="Excel file path")
    args = parser.parse_args(argv)

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        return 1

    wb = load_workbook(str(path), data_only=False)
    findings: List[Dict[str, str]] = []

    # Pre-collect max_row for every sheet (used for cross-sheet ref check)
    sheet_max_rows: Dict[str, int] = {}
    sheet_max_cols: Dict[str, int] = {}
    for sn in wb.sheetnames:
        sw = wb[sn]
        sheet_max_rows[sn] = sw.max_row or 1
        sheet_max_cols[sn] = sw.max_column or 1

    for sname in wb.sheetnames:
        ws = wb[sname]
        max_data_row = sheet_max_rows[sname]
        max_data_col = sheet_max_cols[sname]

        # Collect formulas by column for pattern analysis
        col_formulas: Dict[Tuple[str, str], List[Tuple[int, str]]] = defaultdict(list)

        for row in ws.iter_rows(min_row=1, max_row=max_data_row,
                                min_col=1, max_col=max_data_col):
            for c in row:
                if not is_formula(c.value):
                    continue
                fstr = str(c.value)
                col_letter = get_column_letter(c.column)
                col_formulas[(sname, col_letter)].append((c.row, fstr))

                # --- Out-of-range references ---
                # Check for cross-sheet references: SheetName!A1:A242
                cross_sheet = re.findall(r"([A-Za-z_]\w*?)!([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)", fstr)
                if cross_sheet:
                    for tgt_sheet, c1, r1, c2, r2 in cross_sheet:
                        tgt_rows = sheet_max_rows.get(tgt_sheet, 0)
                        r2_int = int(r2)
                        # Only flag if target sheet exists and range truly exceeds it
                        if tgt_rows > 0 and r2_int > tgt_rows * 3 and r2_int > 100:
                            findings.append({
                                "type": "out_of_range",
                                "cell": cell_ref(sname, c),
                                "detail": f"Range {tgt_sheet}!{c1}{r1}:{c2}{r2} extends far beyond {tgt_sheet} data ({tgt_rows} rows)",
                                "formula": fstr[:100],
                            })
                else:
                    # Same-sheet reference check
                    range_refs = re.findall(r'([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)', fstr)
                    for c1, r1, c2, r2 in range_refs:
                        r1_int, r2_int = int(r1), int(r2)
                        if r2_int > max_data_row * 3 and r2_int > 100:
                            findings.append({
                                "type": "out_of_range",
                                "cell": cell_ref(sname, c),
                                "detail": f"Range {c1}{r1}:{c2}{r2} extends far beyond data ({max_data_row} rows)",
                                "formula": fstr[:100],
                            })

                # --- Header row inclusion ---
                agg_pattern = re.compile(
                    r'(SUM|AVERAGE|AVG|COUNT|COUNTA|MIN|MAX|SUMPRODUCT)\s*\(\s*([A-Z]{1,3})1:',
                    re.IGNORECASE,
                )
                agg_match = agg_pattern.search(fstr)
                if agg_match and c.row > 1:
                    findings.append({
                        "type": "header_included",
                        "cell": cell_ref(sname, c),
                        "detail": f"{agg_match.group(1)}() starts at row 1 (header row)",
                        "formula": fstr[:100],
                    })

                # --- Insufficient aggregate range ---
                small_range = re.compile(
                    r'(SUM|AVERAGE|AVG|COUNT|COUNTA)\s*\(\s*([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)\s*\)',
                    re.IGNORECASE,
                )
                for m in small_range.finditer(fstr):
                    func, _, r1s, _, r2s = m.groups()
                    span = abs(int(r2s) - int(r1s)) + 1
                    if span <= 2:
                        findings.append({
                            "type": "small_aggregate",
                            "cell": cell_ref(sname, c),
                            "detail": f"{func}() covers only {span} cell(s)",
                            "formula": fstr[:100],
                        })

        # --- Inconsistent formula patterns within same column ---
        for (sh, col), entries in col_formulas.items():
            if len(entries) < 3:
                continue
            patterns: Dict[str, List[int]] = defaultdict(list)
            for row_num, fstr in entries:
                norm = re.sub(r'(\$?[A-Z]{1,3}\$?)\d+', r'\1#', fstr)
                patterns[norm].append(row_num)

            if len(patterns) <= 1:
                continue

            dominant_pat = max(patterns, key=lambda k: len(patterns[k]))
            dominant_rows = patterns[dominant_pat]

            for pat, rows in patterns.items():
                if pat == dominant_pat:
                    continue
                if len(rows) <= 2 and len(rows) < len(dominant_rows):
                    for r in rows:
                        orig = next((f for rn, f in entries if rn == r), "?")
                        findings.append({
                            "type": "inconsistent_pattern",
                            "cell": f"{sh}!{col}{r}",
                            "detail": f"Formula differs from {len(dominant_rows)} other rows in column {col}",
                            "formula": orig[:100],
                        })

    wb.close()

    result: Dict[str, Any] = {
        "total_findings": len(findings),
        "by_type": {},
    }
    for f in findings:
        t = f["type"]
        result["by_type"].setdefault(t, 0)
        result["by_type"][t] += 1

    if findings:
        result["findings"] = findings[:60]

    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 1 if findings else 0


# ═══════════════════════════════════════════════════════════════
#  Section 4: inspect — Structure analysis → JSON
# ═══════════════════════════════════════════════════════════════

@cmd("inspect")
def cmd_inspect(argv: Sequence[str]) -> int:
    """Analyse Excel file structure and output JSON."""
    parser = argparse.ArgumentParser(prog="xlsx.py inspect",
                                     description="Analyse file structure → JSON")
    parser.add_argument("file", help="Excel file path")
    parser.add_argument("--pretty", action="store_true", help="Pretty-print JSON")
    args = parser.parse_args(argv)

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        return 1

    wb = load_workbook(str(path), data_only=False, read_only=False)
    sheets_info: List[Dict[str, Any]] = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Extract headers (first row)
        headers: List[Optional[str]] = []
        if max_row > 0 and max_col > 0:
            for cell in ws[1]:
                if cell.value is not None:
                    headers.append(str(cell.value))
                else:
                    headers.append(None)

        # Build data range string
        if max_row > 0 and max_col > 0:
            data_range = f"A1:{get_column_letter(max_col)}{max_row}"
        else:
            data_range = None

        # Count formulas and data rows
        formula_count = 0
        data_rows = 0
        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            has_data = False
            for c in row:
                if c.value is not None:
                    has_data = True
                    if is_formula(c.value):
                        formula_count += 1
            if has_data:
                data_rows += 1

        # Detect charts
        chart_count = 0
        if hasattr(ws, "_charts"):
            chart_count = len(ws._charts)

        sheet_info: Dict[str, Any] = {
            "name": sname,
            "dataRange": data_range,
            "rows": max_row,
            "columns": max_col,
            "dataRows": data_rows,
            "formulaCount": formula_count,
            "chartCount": chart_count,
            "tables": [{
                "headers": [h for h in headers if h is not None]
            }] if headers and any(h is not None for h in headers) else [],
        }
        sheets_info.append(sheet_info)

    wb.close()

    output = {"sheets": sheets_info}
    indent = 2 if args.pretty else None
    print(json.dumps(output, indent=indent, ensure_ascii=False))
    return 0


# ═══════════════════════════════════════════════════════════════
#  Section 5: pivot — PivotTable with optional chart
# ═══════════════════════════════════════════════════════════════

def _aggregate(values: List[float], method: str) -> float:
    """Compute aggregation on a list of numbers."""
    if not values:
        return 0.0
    if method == "sum":
        return sum(values)
    elif method == "count":
        return float(len(values))
    elif method == "average":
        return sum(values) / len(values)
    elif method == "max":
        return max(values)
    elif method == "min":
        return min(values)
    return sum(values)


@cmd("pivot")
def cmd_pivot(argv: Sequence[str]) -> int:
    """Create a PivotTable-like summary with optional chart using openpyxl."""
    parser = argparse.ArgumentParser(prog="xlsx.py pivot",
                                     description="Create PivotTable summary with optional chart")
    parser.add_argument("input", help="Input Excel file")
    parser.add_argument("output", help="Output Excel file")
    parser.add_argument("--source", required=True, help="Source range: 'Sheet!A1:Z100'")
    parser.add_argument("--values", required=True, help="Value fields: 'Revenue:sum,Units:count'")
    parser.add_argument("--rows", default=None, help="Row fields: 'Product,Region'")
    parser.add_argument("--cols", default=None, help="Column fields: 'Quarter'")
    parser.add_argument("--filters", default=None, help="Filter fields: 'Year'")
    parser.add_argument("--location", default="PivotTable!A3", help="Output location: 'Sheet!A3'")
    parser.add_argument("--name", default="PivotTable1", help="PivotTable name")
    parser.add_argument("--style", default="monochrome", choices=["monochrome", "finance"],
                        help="Visual style theme")
    parser.add_argument("--chart", default=None, choices=["bar", "line", "pie"],
                        help="Chart type (optional)")
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(json.dumps({"error": f"Input file not found: {input_path}"}))
        return 1

    # Parse source range
    src_sheet, src_min_col, src_min_row, src_max_col, src_max_row = parse_range(args.source)

    # Parse location
    loc = parse_range(args.location)
    loc_sheet = loc[0]
    loc_start_col = loc[1]
    loc_start_row = loc[2]

    # Parse value fields
    value_fields: List[Tuple[str, str]] = []
    for vspec in args.values.split(","):
        vspec = vspec.strip()
        if ":" in vspec:
            fname, agg = vspec.rsplit(":", 1)
            agg = agg.strip().lower()
            if agg in ("avg", "average"):
                agg = "average"
        else:
            fname = vspec
            agg = "sum"
        value_fields.append((fname.strip(), agg))

    row_fields = [f.strip() for f in args.rows.split(",") if f.strip()] if args.rows else []
    col_fields = [f.strip() for f in args.cols.split(",") if f.strip()] if args.cols else []
    filter_fields = [f.strip() for f in args.filters.split(",") if f.strip()] if args.filters else []

    # Load workbook
    wb = load_workbook(str(input_path))
    if src_sheet not in wb.sheetnames:
        print(json.dumps({"error": f"Source sheet '{src_sheet}' not found"}))
        return 1

    ws_src = wb[src_sheet]

    # Read headers from first row of source
    headers: List[str] = []
    for col_idx in range(src_min_col, src_max_col + 1):
        val = ws_src.cell(row=src_min_row, column=col_idx).value
        headers.append(str(val) if val is not None else f"Col{col_idx}")

    # Build column index map
    col_map = {h: i for i, h in enumerate(headers)}

    # Validate field names
    all_fields = row_fields + col_fields + [vf[0] for vf in value_fields] + filter_fields
    for f in all_fields:
        if f not in col_map:
            print(json.dumps({"error": f"Field '{f}' not found. Available: {headers}"}))
            return 1

    # Read data rows
    data_rows: List[Dict[str, Any]] = []
    for row_idx in range(src_min_row + 1, src_max_row + 1):
        row_data: Dict[str, Any] = {}
        for col_idx in range(src_min_col, src_max_col + 1):
            h = headers[col_idx - src_min_col]
            row_data[h] = ws_src.cell(row=row_idx, column=col_idx).value
        if any(v is not None for v in row_data.values()):
            data_rows.append(row_data)

    # Aggregate data
    def make_key(row: Dict[str, Any], fields: List[str]) -> Tuple[str, ...]:
        return tuple(str(row.get(f, "")) for f in fields)

    group_fields = row_fields + col_fields
    groups: Dict[Tuple[str, ...], Dict[str, List[float]]] = defaultdict(lambda: defaultdict(list))

    for row in data_rows:
        key = make_key(row, group_fields)
        for vname, _ in value_fields:
            val = row.get(vname)
            if isinstance(val, (int, float)):
                groups[key][vname].append(float(val))

    # Create or get output sheet
    if loc_sheet and loc_sheet in wb.sheetnames:
        ws_out = wb[loc_sheet]
    elif loc_sheet:
        ws_out = wb.create_sheet(loc_sheet)
    else:
        ws_out = wb.create_sheet("PivotTable")

    # ---- Styling ----
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # --- Font resolution (mirrors templates/base.py logic) ---
    _platform_hints = {
        "Darwin":  {"PingFang SC", "Hiragino Sans GB"},
        "Windows": {"Microsoft YaHei", "SimHei"},
        "Linux":   {"Noto Sans CJK SC", "WenQuanYi Micro Hei"},
    }
    _cjk_chain = ["PingFang SC", "Microsoft YaHei", "Noto Sans CJK SC",
                   "Hiragino Sans GB", "Source Han Sans SC", "SimHei"]
    _avail = _platform_hints.get(platform.system(), set())
    _font_name = next((f for f in _cjk_chain if f in _avail), _cjk_chain[0])
    _heavy = {"SimHei", "Microsoft YaHei", "PingFang SC", "Noto Sans CJK SC",
              "Source Han Sans SC", "Hiragino Sans GB", "WenQuanYi Micro Hei"}
    _header_bold = _font_name not in _heavy

    if args.style == "finance":
        header_fill_color = "1B2A4A"  # PRIMARY from design token
        alt_row_color = "D6E4F0"      # PRIMARY_LIGHT
    else:  # monochrome
        header_fill_color = "333333"
        alt_row_color = "F5F5F5"

    header_fill = PatternFill(start_color=header_fill_color,
                              end_color=header_fill_color, fill_type="solid")
    header_font = Font(name=_font_name, color="FFFFFF", bold=_header_bold, size=11)
    data_font = Font(name=_font_name, size=11)
    alt_fill = PatternFill(start_color=alt_row_color,
                           end_color=alt_row_color, fill_type="solid")
    border = Border(bottom=Side(style="thin", color="D0D0D0"))

    # Determine if cross-matrix mode (--cols provided)
    use_cross_matrix = len(col_fields) > 0

    if use_cross_matrix:
        # ── Cross-matrix mode: row_fields as rows, col_fields expanded as columns ──
        # Collect unique column dimension values
        col_dim_values: List[str] = []
        seen_col_vals: set = set()
        for row in data_rows:
            cv = str(row.get(col_fields[0], ""))
            if cv not in seen_col_vals:
                seen_col_vals.add(cv)
                col_dim_values.append(cv)
        col_dim_values.sort()

        # Build cross-matrix groups: key = row_fields only, sub-key = col_dim value
        cross_groups: Dict[Tuple[str, ...], Dict[str, Dict[str, List[float]]]] = defaultdict(
            lambda: defaultdict(lambda: defaultdict(list))
        )
        for row in data_rows:
            rkey = make_key(row, row_fields)
            cval = str(row.get(col_fields[0], ""))
            for vname, _ in value_fields:
                val = row.get(vname)
                if isinstance(val, (int, float)):
                    cross_groups[rkey][cval][vname].append(float(val))

        # Build output headers: row_fields + (col_val - agg_name) for each combination
        out_headers: List[str] = list(row_fields)
        for cv in col_dim_values:
            for vname, agg in value_fields:
                if len(value_fields) == 1:
                    out_headers.append(f"{cv}")
                else:
                    out_headers.append(f"{cv} ({vname} {agg})")

        # Write headers
        r = loc_start_row
        for i, h in enumerate(out_headers):
            cell = ws_out.cell(row=r, column=loc_start_col + i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sort row keys
        sorted_row_keys = sorted(cross_groups.keys())

        # Write data rows
        for idx, rkey in enumerate(sorted_row_keys):
            r += 1
            for i, val in enumerate(rkey):
                cell = ws_out.cell(row=r, column=loc_start_col + i, value=val)
                cell.font = data_font
                cell.border = border
                if idx % 2 == 1:
                    cell.fill = alt_fill

            col_offset = len(row_fields)
            for cv in col_dim_values:
                for vname, agg in value_fields:
                    vals = cross_groups[rkey].get(cv, {}).get(vname, [])
                    agg_result = _aggregate(vals, agg) if vals else 0
                    cell = ws_out.cell(row=r, column=loc_start_col + col_offset,
                                       value=round(agg_result, 2))
                    cell.font = data_font
                    cell.border = border
                    cell.number_format = "#,##0.00"
                    if idx % 2 == 1:
                        cell.fill = alt_fill
                    col_offset += 1

        sorted_keys = sorted_row_keys
        total_data_rows_for_chart = len(sorted_row_keys)

    else:
        # ── Flat mode: no --cols, original behavior ──
        out_headers: List[str] = list(row_fields)
        for vname, agg in value_fields:
            out_headers.append(f"{vname} ({agg})")

        # Write headers
        r = loc_start_row
        for i, h in enumerate(out_headers):
            cell = ws_out.cell(row=r, column=loc_start_col + i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sort keys for consistent output
        sorted_keys = sorted(groups.keys())

        # Write data rows
        for idx, key in enumerate(sorted_keys):
            r += 1
            # Row fields
            for i, val in enumerate(key[:len(row_fields)]):
                cell = ws_out.cell(row=r, column=loc_start_col + i, value=val)
                cell.font = data_font
                cell.border = border
                if idx % 2 == 1:
                    cell.fill = alt_fill

            # Value fields
            for i, (vname, agg) in enumerate(value_fields):
                vals = groups[key].get(vname, [])
                agg_result = _aggregate(vals, agg)
                col_offset = len(row_fields) + i
                cell = ws_out.cell(row=r, column=loc_start_col + col_offset,
                                   value=round(agg_result, 2))
                cell.font = data_font
                cell.border = border
                cell.number_format = "#,##0.00"
                if idx % 2 == 1:
                    cell.fill = alt_fill

        total_data_rows_for_chart = len(sorted_keys)

    # Auto-adjust column widths (data-driven, headers wrap if too wide)
    try:
        from templates.base import auto_fit_columns
        auto_fit_columns(ws_out, min_width=10, max_width=28,
                         header_row=loc_start_row, data_start_row=loc_start_row + 1)
    except ImportError:
        # Fallback: old logic
        for i, h in enumerate(out_headers):
            col_letter = get_column_letter(loc_start_col + i)
            ws_out.column_dimensions[col_letter].width = max(len(str(h)) + 4, 14)

    # Hide gridlines
    ws_out.sheet_view.showGridLines = False

    # Add chart if requested
    has_chart = False
    total_data_rows = total_data_rows_for_chart
    if args.chart and total_data_rows > 0:
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        chart_type_map = {
            "bar": BarChart,
            "line": LineChart,
            "pie": PieChart,
        }
        ChartClass = chart_type_map.get(args.chart, BarChart)
        chart = ChartClass()
        chart.title = args.name or "PivotTable Summary"
        chart.style = 10

        data_ref = Reference(
            ws_out,
            min_col=loc_start_col + len(row_fields),
            min_row=loc_start_row,
            max_col=loc_start_col + len(out_headers) - 1,
            max_row=loc_start_row + total_data_rows,
        )
        cats_ref = Reference(
            ws_out,
            min_col=loc_start_col,
            min_row=loc_start_row + 1,
            max_row=loc_start_row + total_data_rows,
        )

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        if isinstance(chart, BarChart):
            chart.type = "col"

        # Pie chart: prevent label overlap with bestFit + leader lines
        if isinstance(chart, PieChart):
            from openpyxl.chart.label import DataLabelList
            chart.dataLabels = DataLabelList()
            chart.dataLabels.dLblPos = 'bestFit'
            chart.dataLabels.showLeaderLines = True
            chart.dataLabels.showCatName = True
            chart.dataLabels.showPercent = True
            chart.dataLabels.showVal = False
            chart.dataLabels.showSerName = False

        # Anchor with enough vertical offset to avoid chart-to-chart overlap
        # ~15 rows per chart height; leave 2 extra rows gap
        chart_row_offset = 17
        chart_anchor = ws_out.cell(
            row=loc_start_row + total_data_rows + 3,
            column=loc_start_col,
        ).coordinate
        ws_out.add_chart(chart, chart_anchor)
        has_chart = True

    wb.save(str(output_path))
    wb.close()

    # [Fix ③] Auto-recalc after pivot if chart was created, so chart data cache is populated
    if has_chart:
        try:
            _run_libreoffice_recalc_best_effort(str(output_path))
        except Exception:
            pass

    print(json.dumps({
        "status": "success",
        "output": str(output_path),
        "pivot_rows": total_data_rows,
        "fields": {
            "rows": row_fields,
            "columns": col_fields,
            "values": [f"{v}:{a}" for v, a in value_fields],
            "filters": filter_fields,
        },
        "chart": args.chart or "none",
    }, indent=2, ensure_ascii=False))
    return 0


# ═══════════════════════════════════════════════════════════════
#  Section 6: chart-verify — Verify chart data content
# ═══════════════════════════════════════════════════════════════

def _check_charts(filepath: str) -> Tuple[List[Dict], List[Dict]]:
    """Core chart verification logic. Returns (ok_charts, empty_charts)."""
    wb = load_workbook(filepath)
    ok_charts: List[Dict[str, str]] = []
    empty_charts: List[Dict[str, str]] = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        if not hasattr(ws, "_charts"):
            continue
        for chart in ws._charts:
            chart_title = "untitled"
            if chart.title:
                try:
                    parts = []
                    for p in chart.title.tx.rich.paragraphs:
                        for run in p.r:
                            if hasattr(run, 't') and run.t:
                                parts.append(run.t)
                    chart_title = "".join(parts) if parts else "untitled"
                except (AttributeError, TypeError):
                    chart_title = "untitled"
            has_data = False

            for series in (chart.series if hasattr(chart, "series") else []):
                if hasattr(series, "val") and series.val:
                    ref = series.val
                    if hasattr(ref, "numRef") and ref.numRef:
                        cache = ref.numRef.numCache if hasattr(ref.numRef, "numCache") else None
                        if cache and hasattr(cache, "ptCount") and cache.ptCount and cache.ptCount > 0:
                            has_data = True
                            break
                    if hasattr(ref, "numLit") and ref.numLit:
                        has_data = True
                        break

            entry = {"sheet": sname, "title": chart_title}
            if has_data:
                ok_charts.append(entry)
            else:
                empty_charts.append(entry)

    wb.close()
    return ok_charts, empty_charts


@cmd("chart-verify")
def cmd_chart_verify(argv: Sequence[str]) -> int:
    """Verify all charts have actual data content.
    [Fix ④] Automatically recalc first if charts appear empty, then re-check.
    """
    parser = argparse.ArgumentParser(prog="xlsx.py chart-verify",
                                     description="Verify all charts have actual data")
    parser.add_argument("file", help="Excel file path")
    parser.add_argument("--no-auto-recalc", action="store_true",
                        help="Disable automatic recalc before checking")
    args = parser.parse_args(argv)

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        return 1

    ok_charts, empty_charts = _check_charts(str(path))

    # [Fix ④] If there are empty charts and auto-recalc is enabled, try recalc first
    if empty_charts and not args.no_auto_recalc:
        try:
            _run_libreoffice_recalc_best_effort(str(path))
            # Re-check after recalc
            ok_charts, empty_charts = _check_charts(str(path))
        except Exception:
            pass

    total_charts = len(ok_charts) + len(empty_charts)

    result: Dict[str, Any] = {
        "total_charts": total_charts,
        "charts_with_data": len(ok_charts),
        "empty_charts": len(empty_charts),
    }
    if empty_charts:
        result["empty"] = empty_charts
    if ok_charts:
        result["ok"] = ok_charts

    print(json.dumps(result, indent=2, ensure_ascii=False))

    if empty_charts:
        return 1
    if total_charts == 0:
        print("No charts found in workbook.", file=sys.stderr)
        return 0
    return 0


# ═══════════════════════════════════════════════════════════════
#  Section 7: validate — Structural validation
# ═══════════════════════════════════════════════════════════════

@cmd("validate")
def cmd_validate(argv: Sequence[str]) -> int:
    """Structural validation: forbidden functions, formula hygiene, schema basics."""
    parser = argparse.ArgumentParser(prog="xlsx.py validate",
                                     description="Structural validation (forbidden funcs, schema)")
    parser.add_argument("file", help="Excel file path")
    args = parser.parse_args(argv)

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        return 1

    wb = load_workbook(str(path), data_only=False)
    issues: List[Dict[str, str]] = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for c in row:
                if not is_formula(c.value):
                    continue
                fstr = str(c.value).upper()

                # Check for forbidden functions
                for func in FORBIDDEN_FUNCTIONS:
                    pattern = re.compile(rf'\b{func}\s*\(', re.IGNORECASE)
                    if pattern.search(fstr):
                        issues.append({
                            "type": "forbidden_function",
                            "cell": cell_ref(sname, c),
                            "function": func,
                            "detail": f"{func}() is not supported in Excel 2019 and earlier",
                            "formula": str(c.value)[:100],
                        })

                # Check for text accidentally treated as formula
                raw = str(c.value)
                is_formula_cell = getattr(c, "data_type", None) == "f"
                if (raw.startswith("=") and not is_formula_cell
                        and not any(ch in raw for ch in "+-*/()&,!:$")):
                    issues.append({
                        "type": "text_as_formula",
                        "cell": cell_ref(sname, c),
                        "detail": "Text starts with '=' — may be misinterpreted as formula",
                        "value": raw[:80],
                    })

                # [Fix ①] Heuristic: data_type=='f' but content has no valid formula elements
                if is_formula_cell and raw.startswith("="):
                    body = raw[1:]  # strip leading =
                    body_stripped = body.strip()
                    if body_stripped.startswith('"'):
                        pass  # starts with a quoted string, likely intentional
                    elif not _VALID_FORMULA_PATTERN.search(body):
                        issues.append({
                            "type": "text_as_formula",
                            "cell": cell_ref(sname, c),
                            "detail": "Cell stored as formula but contains no function calls or cell references — likely text starting with '='",
                            "value": raw[:80],
                        })

                # [Fix ②] Check for external file references in formulas
                if _EXT_FILE_REF_PATTERN.search(fstr):
                    ext_matches = _EXT_FILE_REF_PATTERN.findall(fstr)
                    for ext_file, _ in ext_matches:
                        issues.append({
                            "type": "external_file_ref",
                            "cell": cell_ref(sname, c),
                            "detail": f"Formula references external file: {ext_file}",
                            "formula": str(c.value)[:100],
                        })

    # Check for absolute paths / local file references in .rels
    try:
        with zipfile.ZipFile(str(path), "r") as zf:
            for name in zf.namelist():
                if name.endswith(".rels"):
                    content = zf.read(name).decode("utf-8", errors="ignore")
                    if re.search(r'Target="[A-Z]:\\', content):
                        issues.append({
                            "type": "absolute_path",
                            "file": name,
                            "detail": "Absolute Windows path in .rels file — causes Excel crash",
                        })
                    if "file:///" in content.lower():
                        issues.append({
                            "type": "local_file_ref",
                            "file": name,
                            "detail": "Local file:// reference in .rels — may cause security warning",
                        })
    except Exception:
        pass  # zipfile inspection is best-effort

    wb.close()

    by_type: Dict[str, int] = defaultdict(int)
    for iss in issues:
        by_type[iss["type"]] += 1

    result: Dict[str, Any] = {
        "status": "passed" if not issues else "failed",
        "total_issues": len(issues),
        "by_type": dict(by_type),
    }
    if issues:
        result["issues"] = issues[:50]

    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0 if not issues else 1


# ═══════════════════════════════════════════════════════════════
#  Section 8: CLI entry point
# ═══════════════════════════════════════════════════════════════

_HELP_TEXT = """\
xlsx.py — Unified Excel Quality Assurance & Manipulation Tool

Usage: python3 xlsx.py <command> [args...]

Commands:
  recalc       <xlsx> [timeout]                  Recalculate formulas via LibreOffice + error scan
  audit      <xlsx>                            Formula error + zero-value + implicit array detection
  scan     <xlsx>                            Reference anomaly detection
  inspect      <xlsx> [--pretty]                 Structure analysis → JSON
  pivot        <in> <out> --source ... [options] PivotTable with optional chart
  chart-verify <xlsx> [--no-auto-recalc]         Verify chart data content
  validate     <xlsx>                            Structural validation

Run 'python3 xlsx.py <command> --help' for command-specific options.
"""


def main() -> int:
    if len(sys.argv) < 2 or sys.argv[1] in ("--help", "-h"):
        print(_HELP_TEXT)
        return 0

    command = sys.argv[1]
    rest = sys.argv[2:]

    handler = _COMMANDS.get(command)
    if handler is None:
        print(f"Unknown command: {command}\n", file=sys.stderr)
        print(_HELP_TEXT, file=sys.stderr)
        return 1

    return handler(rest)


if __name__ == "__main__":
    sys.exit(main())
