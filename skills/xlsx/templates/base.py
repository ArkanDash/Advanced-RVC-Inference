"""
xlsx skill — Base Template
===========================
Single source of truth for design tokens, font resolution, and style factories.
All scene/engine code MUST import from here. Never hardcode colors, fonts, or styles.

Usage:
    from templates.base import *

    # To switch palette based on user prompt (call BEFORE creating styles):
    use_palette("帮我做一个温暖的销售月报")  # Chinese prompt example
    # → All color tokens and style factories now use 'warm' palette.

    # Or manually:
    use_palette_explicit("warm")
"""

import platform
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from copy import copy


# ============================================================
# §1  Font Resolution (cross-platform fallback chain)
# ============================================================

def _resolve_font(candidates: list) -> str:
    """Return the first font name likely available on this OS."""
    system = platform.system()
    _platform_hints = {
        "Darwin":  {"PingFang SC", "Hiragino Sans GB", ".AppleSystemUIFont"},
        "Windows": {"Microsoft YaHei", "SimHei", "SimSun"},
        "Linux":   {"Noto Sans CJK SC", "WenQuanYi Micro Hei", "Source Han Sans SC"},
    }
    available = _platform_hints.get(system, set())
    for name in candidates:
        if name in available:
            return name
    return candidates[0]


# CJK sans-serif fallback chain
CJK_BODY_CHAIN = [
    "PingFang SC",         # macOS
    "Microsoft YaHei",     # Windows
    "Noto Sans CJK SC",    # Linux / Android
    "Hiragino Sans GB",    # macOS alt
    "Source Han Sans SC",  # Adobe cross-platform
    "SimHei",              # classic fallback
]

# Latin serif (for formal reports)
LATIN_BODY_CHAIN = [
    "Times New Roman",
    "Georgia",
    "serif",
]

FONT_CJK   = _resolve_font(CJK_BODY_CHAIN)
FONT_LATIN  = _resolve_font(LATIN_BODY_CHAIN)

# Primary font — CJK font covers ASCII too
FONT_NAME = FONT_CJK

# Bold strategy: heavy-stroke fonts should NOT be bolded
_HEAVY_FONTS = {
    "SimHei", "Microsoft YaHei", "PingFang SC",
    "Noto Sans CJK SC", "Source Han Sans SC",
    "Hiragino Sans GB", "WenQuanYi Micro Hei",
}
HEADER_BOLD = FONT_NAME not in _HEAVY_FONTS


# ============================================================
# §2  Color Tokens (Three-Color Rule)
# ============================================================

# --- Primary (deep blue — professional default) ---
PRIMARY       = "1B2A4A"
PRIMARY_LIGHT = "D6E4F0"
SECONDARY     = PRIMARY_LIGHT   # derived from primary

# --- Accent (semantic, on-demand) ---
ACCENT_POSITIVE = "1B7D46"      # growth, done, pass   (deep green)
ACCENT_NEGATIVE = "C0392B"      # decline, overdue     (deep red)
ACCENT_WARNING  = "D4820A"      # at-risk, watch       (deep amber)

# --- Neutral (warm gray) ---
NEUTRAL_900 = "37352F"          # body text
NEUTRAL_600 = "8C8A84"          # caption, secondary text
NEUTRAL_200 = "E9E9E8"          # borders, dividers
NEUTRAL_100 = "F7F7F5"          # alternating row fill (odd)
NEUTRAL_50  = "FAFAF9"          # ultra-light bg (optional)
NEUTRAL_0   = "FFFFFF"          # white (even rows)

# --- Header text color (overridable by palette) ---
HEADER_TEXT = "FFFFFF"

# --- Chart palette (max 5 colors) ---
CHART_COLORS = [PRIMARY, ACCENT_POSITIVE, ACCENT_WARNING, ACCENT_NEGATIVE, NEUTRAL_600]

# --- Conditional formatting fills ---
CF_POSITIVE_FILL = PatternFill(bgColor="E8F5E9")
CF_POSITIVE_FONT = Font(color=ACCENT_POSITIVE)
CF_NEGATIVE_FILL = PatternFill(bgColor="FDEDEC")
CF_NEGATIVE_FONT = Font(color=ACCENT_NEGATIVE)
CF_WARNING_FILL  = PatternFill(bgColor="FEF9E7")
CF_WARNING_FONT  = Font(color=ACCENT_WARNING)

# --- Active style (for debugging/logging) ---
_ACTIVE_STYLE = "professional"


# ============================================================
# §2.1  Palette Integration
# ============================================================

def use_palette(prompt: str):
    """
    Auto-detect style from user prompt and switch all color tokens.
    Call this BEFORE creating any styles/cells.

    Three-step matching:
      1. Explicit style keywords → direct match
      2. Scene/content keywords → infer style
      3. No match → professional (safe default)

    Example:
        use_palette("帮我做一个温暖的销售月报")  # Chinese prompt example
        # → 'warm' palette applied
    """
    from templates.palettes import resolve_palette_with_info
    palette, style = resolve_palette_with_info(prompt)
    _apply(palette, style)


def use_palette_explicit(style: str = "professional"):
    """
    Manually select a palette by style name.
    Available: professional, warm, elegant, creative, muji, aesop,
               kinfolk, celine, bottega, chanel, bloomberg, original_blue

    Example:
        use_palette_explicit("warm")
    """
    from templates.palettes import get_palette
    palette = get_palette(style)
    _apply(palette, style)


def get_active_style() -> str:
    """Return the currently active style name."""
    return _ACTIVE_STYLE


def _apply(palette: dict, style: str):
    """Internal: apply a palette dict to all module-level color tokens."""
    global PRIMARY, PRIMARY_LIGHT, SECONDARY
    global ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING
    global NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0
    global CHART_COLORS, HEADER_TEXT
    global CF_POSITIVE_FILL, CF_POSITIVE_FONT
    global CF_NEGATIVE_FILL, CF_NEGATIVE_FONT
    global CF_WARNING_FILL, CF_WARNING_FONT
    global _ACTIVE_STYLE

    PRIMARY       = palette["PRIMARY"]
    PRIMARY_LIGHT = palette["PRIMARY_LIGHT"]
    SECONDARY     = palette["SECONDARY"]
    ACCENT_POSITIVE = palette["ACCENT_POSITIVE"]
    ACCENT_NEGATIVE = palette["ACCENT_NEGATIVE"]
    ACCENT_WARNING  = palette["ACCENT_WARNING"]
    NEUTRAL_900 = palette["NEUTRAL_900"]
    NEUTRAL_600 = palette["NEUTRAL_600"]
    NEUTRAL_200 = palette["NEUTRAL_200"]
    NEUTRAL_100 = palette["NEUTRAL_100"]
    NEUTRAL_50  = palette["NEUTRAL_50"]
    NEUTRAL_0   = palette["NEUTRAL_0"]
    HEADER_TEXT  = palette.get("HEADER_TEXT", "FFFFFF")
    CHART_COLORS = palette["CHART_COLORS"]

    # Rebuild conditional formatting fills/fonts with new accent colors
    CF_POSITIVE_FILL = PatternFill(bgColor=palette.get("CF_POSITIVE_BG", "E8F5E9"))
    CF_POSITIVE_FONT = Font(color=ACCENT_POSITIVE)
    CF_NEGATIVE_FILL = PatternFill(bgColor=palette.get("CF_NEGATIVE_BG", "FDEDEC"))
    CF_NEGATIVE_FONT = Font(color=ACCENT_NEGATIVE)
    CF_WARNING_FILL  = PatternFill(bgColor=palette.get("CF_WARNING_BG", "FEF9E7"))
    CF_WARNING_FONT  = Font(color=ACCENT_WARNING)

    _ACTIVE_STYLE = style


# ============================================================
# §3  Column Width Map
# ============================================================

COLUMN_WIDTHS = {
    "margin":      3,     # A col whitespace
    "id_short":    8,     # #, ID
    "name_cn":    16,     # Chinese name (2-4 chars)
    "name_en":    22,     # English name
    "description": 32,    # long text
    "number":     14,     # currency, amount
    "percentage": 12,     # %
    "date":       14,     # YYYY-MM-DD
    "status":     12,     # short label
}


# ============================================================
# §4  Number Formats
# ============================================================

FORMATS = {
    "integer":      "#,##0",
    "decimal_1":    "#,##0.0",
    "decimal_2":    "#,##0.00",
    "percentage":   "0.0%",
    "currency_cny": "¥#,##0.00",
    "currency_usd": "$#,##0.00",
    "date":         "YYYY-MM-DD",
}


# ============================================================
# §5  Style Factories
# ============================================================

def font_title():
    """16pt title font — left-aligned, no fill."""
    return Font(name=FONT_NAME, size=16, bold=HEADER_BOLD, color=PRIMARY)

def font_header():
    """11pt header font — text color on primary background."""
    return Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=HEADER_TEXT)

def font_subheader():
    """11pt sub-header — primary color text."""
    return Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PRIMARY)

def font_body():
    """11pt body text."""
    return Font(name=FONT_NAME, size=11, color=NEUTRAL_900)

def font_caption():
    """9pt caption / footnote."""
    return Font(name=FONT_NAME, size=9, color=NEUTRAL_600)

def font_kpi():
    """22pt big KPI number."""
    return Font(name=FONT_NAME, size=22, bold=HEADER_BOLD, color=PRIMARY)

def font_kpi_label():
    """9pt KPI label."""
    return Font(name=FONT_NAME, size=9, color=NEUTRAL_600)


def make_chart_title(text, size_pt=12, bold=True, axis=False, max_line_chars=6):
    """
    Build a chart Title with font baked into <tx><rich><defRPr>/<rPr>.
    Ensures WPS and Office render identical font name and size.
    Uses FONT_NAME and HEADER_BOLD from §1 — no hardcoded font names.

    Args:
        axis: If True, set bodyPr rot=-5400000 (rotate -90°) for Y-axis titles.
        max_line_chars: For axis titles, auto-insert line breaks (\n) when text
              exceeds this length. Breaks at parentheses boundaries.
              The text stays in ONE run inside ONE paragraph — this prevents
              WPS/Office from creating separate overlapping text boxes.
              Set to 0 or None to disable.

    Key insight: Multiple <p> paragraphs in axis titles cause WPS to render
    them as stacked overlapping text boxes. Instead, we use a SINGLE <r> run
    with \\n line breaks inside the text, which both Office and WPS render
    as line breaks within the same text box.
    """
    from openpyxl.chart.title import Title
    from openpyxl.chart.text import Text, RichText
    from openpyxl.drawing.text import (
        Paragraph, ParagraphProperties, CharacterProperties,
        Font as DrawingFont, RichTextProperties, RegularTextRun,
        LineBreak,
    )
    from copy import deepcopy
    import re

    rpr = CharacterProperties(
        latin=DrawingFont(typeface=FONT_NAME),
        ea=DrawingFont(typeface=FONT_NAME),
        sz=int(size_pt * 100),
        b=bold if HEADER_BOLD else False,
    )

    def _insert_breaks(text, max_chars):
        """Insert \\n before parentheses when text exceeds max_chars."""
        if not max_chars or len(text) <= max_chars:
            return text
        # Insert \n before '(' or '（'
        result = re.sub(r'(?=[（(])', '\n', text, count=1)
        return result

    # For axis titles, insert line breaks to prevent overlap
    display_text = text
    if axis and max_line_chars:
        display_text = _insert_breaks(text, max_line_chars)

    # Single paragraph, single run with \n inside the text.
    # Both Office and WPS render \n as line breaks within one text box.
    # Do NOT use multiple <p> paragraphs — WPS renders them as separate
    # overlapping text boxes on axis titles.
    run = RegularTextRun(rPr=deepcopy(rpr), t=display_text)

    inner_body = RichTextProperties(rot=-5400000) if axis else RichTextProperties()
    para = Paragraph(
        pPr=ParagraphProperties(defRPr=deepcopy(rpr)),
        r=[run],
    )
    rich = RichText(bodyPr=inner_body, p=[para])

    # Outer txPr: Office reads rotation from here for axis titles
    if axis:
        outer_body = RichTextProperties(rot=-5400000)
        txPr = RichText(
            bodyPr=outer_body,
            p=[Paragraph(pPr=ParagraphProperties(defRPr=deepcopy(rpr)))],
        )
        return Title(tx=Text(rich=rich), txPr=txPr)
    return Title(tx=Text(rich=rich))


def fill_header():
    return PatternFill("solid", fgColor=PRIMARY)

def fill_total():
    return PatternFill("solid", fgColor=SECONDARY)

def fill_data_row(row_index: int):
    """Alternating row: even=white, odd=warm-white."""
    color = NEUTRAL_0 if row_index % 2 == 0 else NEUTRAL_100
    return PatternFill("solid", fgColor=color)


def border_header():
    """Thin bottom border under header row."""
    return Border(bottom=Side(style="thin", color=NEUTRAL_200))

def border_total():
    """Medium top border above totals row."""
    return Border(top=Side(style="medium", color=NEUTRAL_200))


def align_title():
    return Alignment(horizontal="left", vertical="center")

def align_header():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def align_number():
    return Alignment(horizontal="right", vertical="center")

def align_text():
    return Alignment(horizontal="left", vertical="center")

def align_date():
    return Alignment(horizontal="center", vertical="center")


# ============================================================
# §6  Sheet Setup Helpers
# ============================================================

ROW_HEIGHTS = {
    "margin":   15,   # row 1 top whitespace
    "title":    32,   # row 2
    "spacer":    8,   # row 3
    "header":   28,   # row 4
    "data":     22,   # data rows
    "total":    26,   # totals row
}


def setup_sheet(ws, title: str = None, last_col: int = None):
    """
    Apply standard sheet setup:
      - hide grid lines
      - set margin column A width
      - set row 1/2/3 heights
      - optionally write & style title at B2
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = COLUMN_WIDTHS["margin"]
    ws.row_dimensions[1].height = ROW_HEIGHTS["margin"]
    ws.row_dimensions[2].height = ROW_HEIGHTS["title"]
    ws.row_dimensions[3].height = ROW_HEIGHTS["spacer"]

    if title and last_col:
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
        cell = ws.cell(row=2, column=2, value=title)
        cell.font = font_title()
        cell.alignment = align_title()


def style_header_row(ws, row_num: int, col_start: int, col_end: int):
    """Apply header style to a row range."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill_header()
        cell.font = font_header()
        cell.alignment = align_header()
        cell.border = border_header()
    ws.row_dimensions[row_num].height = ROW_HEIGHTS["header"]


def style_data_row(ws, row_num: int, col_start: int, col_end: int, row_index: int):
    """Apply data row style (alternating fill)."""
    fill = fill_data_row(row_index)
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font_body()
    ws.row_dimensions[row_num].height = ROW_HEIGHTS["data"]


def style_total_row(ws, row_num: int, col_start: int, col_end: int):
    """Apply totals row style."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill_total()
        cell.font = font_subheader()
        cell.border = border_total()
    ws.row_dimensions[row_num].height = ROW_HEIGHTS["total"]


# ============================================================
# §6.1  Chart Factory Functions
# ============================================================

def create_bar_chart(chart_type="col", grouping="clustered", gap_width=80,
                     overlap=100, style=10, width=18, height=10, **kwargs):
    """
    Create a BarChart with sane defaults that prevent the "thin bar" / offset bug.

    Key fixes baked in:
      - gapWidth=80 (default 150 → bars too thin)
      - overlap=100 (bars fill their slot, no empty gap for line series)

    Returns an openpyxl BarChart ready for add_data / set_categories.
    """
    from openpyxl.chart import BarChart
    chart = BarChart()
    chart.type = chart_type
    chart.grouping = grouping
    chart.gapWidth = gap_width
    chart.overlap = overlap
    chart.style = style
    chart.width = width
    chart.height = height
    return chart


def create_line_chart(style=10, width=18, height=11, **kwargs):
    """Create a LineChart with standard defaults."""
    from openpyxl.chart import LineChart
    chart = LineChart()
    chart.style = style
    chart.width = width
    chart.height = height
    return chart


def create_pie_chart(style=10, width=14, height=10, **kwargs):
    """Create a PieChart with standard defaults."""
    from openpyxl.chart import PieChart
    chart = PieChart()
    chart.style = style
    chart.width = width
    chart.height = height
    return chart


def setup_chart_titles(chart, title=None, y_title=None, x_title=None,
                       title_size=12, axis_size=10):
    """
    Set chart title and axis titles using make_chart_title() for
    cross-platform font consistency (Office + WPS).

    This is the ONLY correct way to set chart titles. Never do:
        chart.title = "some string"        # ← WRONG
        chart.y_axis.title = "some string" # ← WRONG

    Args:
        chart: openpyxl chart object
        title: main chart title (optional)
        y_title: Y-axis title (optional, auto-rotated -90°)
        x_title: X-axis title (optional)
        title_size: font size for main title (default 12)
        axis_size: font size for axis titles (default 10)
    """
    if title is not None:
        chart.title = make_chart_title(title, size_pt=title_size, bold=True)
    if y_title is not None:
        chart.y_axis.title = make_chart_title(y_title, size_pt=axis_size, bold=False, axis=True)
    if x_title is not None:
        chart.x_axis.title = make_chart_title(x_title, size_pt=axis_size, bold=False)


def apply_chart_colors(chart, colors=None):
    """
    Apply palette colors to all series in a chart.
    Call AFTER add_data().

    Args:
        chart: openpyxl chart object (BarChart, LineChart, etc.)
        colors: list of hex color strings (default: CHART_COLORS)
    """
    if colors is None:
        colors = CHART_COLORS
    for i, series in enumerate(chart.series):
        color_hex = colors[i % len(colors)]
        series.graphicalProperties.solidFill = color_hex
        # For line charts, also set line color
        if hasattr(series.graphicalProperties, 'line') and series.graphicalProperties.line is not None:
            series.graphicalProperties.line.solidFill = color_hex


def apply_pie_colors(chart, count, colors=None):
    """
    Apply palette colors to pie chart data points.
    Call AFTER add_data().

    Args:
        chart: openpyxl PieChart
        count: number of data points (slices)
        colors: list of hex color strings (default: CHART_COLORS)
    """
    from openpyxl.chart.series import DataPoint
    if colors is None:
        colors = CHART_COLORS
    for idx in range(count):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = colors[idx % len(colors)]
        chart.series[0].data_points.append(pt)


# ============================================================
# §7  Utility Functions
# ============================================================

def normalize_cell_value(value):
    """Normalize cell values: convert invisible whitespace variants to None."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip().replace("\xa0", "").replace("\u200b", "")
        if stripped == "":
            return None
    return value


def copy_style(source_cell, target_cell):
    """Copy all styling from source to target cell."""
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format


def auto_fit_columns(ws, min_width=8, max_width=28, header_row=None, data_start_row=None):
    """
    Auto-fit column widths based on DATA content (not header).
    Headers that exceed the computed width get wrap_text=True instead of stretching the column.

    Args:
        ws: worksheet
        min_width: minimum column width (default 8)
        max_width: maximum column width (default 28)
        header_row: row number of the header (auto-detected if None)
        data_start_row: first data row (auto-detected as header_row + 1 if None)
    """
    import unicodedata

    def _display_width(text):
        """Estimate display width: CJK chars count as ~1.7, others as 1."""
        if text is None:
            return 0
        s = str(text)
        w = 0
        for ch in s:
            if unicodedata.east_asian_width(ch) in ('W', 'F'):
                w += 1.7
            else:
                w += 1
        return w

    # Auto-detect header row: first row with data starting from column B
    if header_row is None:
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if val is not None:
                header_row = row
                break
        if header_row is None:
            return

    if data_start_row is None:
        data_start_row = header_row + 1

    for col_cells in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=data_start_row, max_row=ws.max_row):
        if not col_cells:
            continue
        col_letter = col_cells[0].column_letter

        # Skip margin column A
        if col_letter == 'A':
            continue

        # Width based on data content only
        max_data_w = max((_display_width(c.value) for c in col_cells), default=0)
        width = min(max_width, max(min_width, max_data_w + 2))
        ws.column_dimensions[col_letter].width = width

        # If header text is wider than computed column width, wrap it
        header_cell = ws.cell(row=header_row, column=col_cells[0].column)
        header_w = _display_width(header_cell.value)
        if header_w > width:
            current_align = header_cell.alignment
            header_cell.alignment = Alignment(
                horizontal=current_align.horizontal or "center",
                vertical=current_align.vertical or "center",
                wrap_text=True,
            )
